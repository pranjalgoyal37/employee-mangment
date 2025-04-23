from flask import Blueprint, request, jsonify
import os
import base64
import numpy as np
import cv2
import datetime
from openpyxl import Workbook, load_workbook
import face_recognition

from models.user_model import users_collection, tasks_collection
from models.attendance_model import attendance_collection,leave_collection
attendance_bp = Blueprint('attendance', __name__)

# Folder for registered face images
REGISTER_FOLDER = "./register_face"

if not os.path.exists(REGISTER_FOLDER):
    os.makedirs(REGISTER_FOLDER)




# Load known faces
def load_known_faces():
    path = REGISTER_FOLDER
    encodings = []
    names = []

    if not os.path.exists(path):
        return encodings, names

    for filename in os.listdir(path):
        if filename.endswith(('.jpg', '.jpeg', '.png')):
            filepath = os.path.join(path, filename)
            img = face_recognition.load_image_file(filepath)
            enc = face_recognition.face_encodings(img)
            if enc:
                encodings.append(enc[0])
                name = os.path.splitext(filename)[0].rsplit("_", 1)[0]
                names.append(name)

    return encodings, names

@attendance_bp.route('/register', methods=['POST'])
def register():
    print("🛠️ You are in register route")
    if request.method == "OPTIONS":
        print("✅ Handling preflight OPTIONS request")
        return '', 200

    try:
        data = request.get_json()
        if not data:
            print("❗ No JSON received")
            return jsonify({'message': 'Invalid JSON'}), 400

        name = data.get('name')
        images = data.get('images', [])

        print("👉 Received name:", name)
        print("👉 Number of images:", len(images))

        if not name or len(images) != 3:
            return jsonify({'message': 'Name or 3 images missing'}), 400

        for i, image_data in enumerate(images):
            try:
                if ',' not in image_data:
                    return jsonify({'message': f'Image {i+1} is not base64 encoded properly'}), 400

                header, encoded = image_data.split(',', 1)
                img_bytes = base64.b64decode(encoded)

                filename = os.path.join(REGISTER_FOLDER, f"{name}_{i+1}.jpg")
                with open(filename, 'wb') as f:
                    f.write(img_bytes)

                print(f"✅ Saved: {filename}")

            except Exception as image_error:
                print(f"❌ Error processing image {i+1}:", image_error)
                return jsonify({'message': f'Error saving image {i+1}', 'error': str(image_error)}), 500

        return jsonify({'message': 'Face registered with 3 images!'})

    except Exception as e:
        print("🔥 Register Error:", e)
        return jsonify({'message': 'Internal server error', 'error': str(e)}), 500


# Mark Attendance via Face Recognition
@attendance_bp.route('/mark_attendance', methods=['POST'])
def mark_attendance():
    try:
        data = request.get_json()
        image_data = data['image']
        header, encoded = image_data.split(",", 1)
        img_bytes = base64.b64decode(encoded)
        np_arr = np.frombuffer(img_bytes, np.uint8)
        frame = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)

        encodings, names = load_known_faces()
        if not encodings:
            return jsonify({"status": "error", "message": "No known faces found"})

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        face_locations = face_recognition.face_locations(rgb_frame)
        face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)

        if not face_encodings:
            return jsonify({"status": "error", "message": "No face detected"})

        now = datetime.datetime.now()
        today_str = now.strftime("%d-%m-%y")
        filename = f"{now.month}.xlsx"
        book = load_workbook(filename) if os.path.exists(filename) else Workbook()
        sheet = book.active
        marked = []

        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(encodings, face_encoding)
            face_distances = face_recognition.face_distance(encodings, face_encoding)

            if True in matches:
                best_index = np.argmin(face_distances)
                if matches[best_index]:
                    name = names[best_index]

                    # Excel marking
                    if sheet.cell(row=1, column=now.day).value is None:
                        sheet.cell(row=1, column=now.day).value = today_str
                    sheet.cell(row=best_index+2, column=1).value = name
                    sheet.cell(row=best_index+2, column=now.day).value = "Present"

                    # MongoDB entry
                    
                    existing = attendance_collection.find_one({"user_name":name ,"date": today_str})
                    print(name,today_str,existing)

                    if existing:
                        return jsonify({"status": "duplicate", "marked": marked})

                    if not  existing:
                        attendance_collection.insert_one({
                            "user_name": name,
                            "date": today_str,
                            "clockIn" : now.strftime("%I:%M:%S %p"),
                            "clockOut": None,
                            "status": "Present"
                        })
                        print(f"user is successfully marked as present")
                    marked.append(name)

        book.save(filename)
        return jsonify({"status": "success", "marked": marked})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500



@attendance_bp.route('/attendance', methods=['GET'])
def get_attendance():
    records = list(attendance_collection.find({}, {"_id": 0}))  # or use {"_id": {"$toString": "$_id"}} to keep id
    return jsonify(records)

@attendance_bp.route('/update-attendance/<string:user_name>', methods=['POST'])
def update_attendance(user_name):
    now = datetime.datetime.now()
    clock_out_time = now.strftime("%I:%M:%S %p")  # 12-hour format with AM/PM
    print(user_name.lower(
        
    ))
    result = attendance_collection.update_one(
        {"user_name": user_name.lower()},
        {"$set": {"clockOut": clock_out_time}}
    )

    if result.modified_count == 1:
        return jsonify({"message": "Attendance updated", "clockOut": clock_out_time}), 200
    else:
        return jsonify({"error": "Failed to update attendance"}), 400

#  leave routes
@attendance_bp.route("/leave-request", methods=["POST"])
def leave_request():
    data = request.json
    if not data:
        return jsonify({"message": "Invalid request"}), 400

    leave_collection.insert_one({
        "start_date": data.get("start_date"),
        "end_date": data.get("end_date"),
        "leave_type": data.get("leave_type"),
        "description": data.get("description"),
        "user_name" :data.get("user_name")
    })

    return jsonify({"message": "Leave request submitted successfully"})


@attendance_bp.route('/leaves', methods=['GET'])
def get_leaves():
    leaves = list(leave_collection.find({}, {'_id': 0}))
    return jsonify(leaves)





